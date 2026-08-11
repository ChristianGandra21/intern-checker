from __future__ import annotations

import csv
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import JobCandidate

HEADERS = [
    "title",
    "company",
    "location",
    "work_mode",
    "source",
    "source_url",
    "published_at",
    "discovered_at",
    "score",
    "display_tier",
    "target_fit",
    "location_fit",
    "display_reasons",
    "score_reasons",
]


def _row(job: JobCandidate) -> list:
    return [
        job.title,
        job.company,
        job.location,
        job.work_mode,
        job.source,
        str(job.source_url),
        job.published_at.isoformat() if job.published_at else "",
        job.discovered_at.isoformat(),
        job.score,
        job.display_tier,
        job.target_fit,
        job.location_fit,
        "; ".join(job.display_reasons),
        "; ".join(job.score_reasons),
    ]


def export_jobs(
    jobs: list[JobCandidate], output_dir: str | Path, filename_prefix: str = "vagas"
) -> tuple[Path, Path]:
    destination = Path(output_dir)
    destination.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now(UTC).strftime("%Y-%m-%d")
    csv_path = destination / f"{filename_prefix}-{stamp}.csv"
    xlsx_path = destination / f"{filename_prefix}-{stamp}.xlsx"

    with csv_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file)
        writer.writerow(HEADERS)
        writer.writerows(_row(job) for job in jobs)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Vagas"
    sheet.append(HEADERS)
    for job in jobs:
        sheet.append(_row(job))
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0E6B4F")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [42, 24, 22, 14, 15, 52, 22, 22, 9, 14, 14, 14, 50, 60]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    workbook.save(xlsx_path)
    return csv_path, xlsx_path
